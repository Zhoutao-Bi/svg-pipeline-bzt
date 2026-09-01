#!/usr/bin/env python3
"""Score dynamic serial outputs with the shared 29-part GT matcher."""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


TARGET_TYPES = {"孔", "柱", "槽", "倒角"}
ASSEMBLY_ROLE = "装配特征"
GT_DIMENSION_CORRECTIONS = {"hard_15": {"尺寸X": 32.42, "尺寸Y": 249.91}}


def number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    return float(match.group()) if match else None


def normalize_gt_type(label) -> str | None:
    text = str(label or "")
    if "孔" in text:
        return "孔"
    if "槽" in text:
        return "槽"
    if "倒角" in text or "圆角" in text:
        return "倒角"
    if any(word in text for word in ("柱", "凸台", "外环", "主直径", "收细段", "半圆")):
        return "柱"
    return None


def load_ground_truth(path: Path) -> dict[str, dict]:
    worksheet = load_workbook(path, read_only=True, data_only=True)["Sheet2"]
    output = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        features = []
        for index in range(5, len(row), 5):
            if row[index] is None:
                continue
            features.append({
                "特征类型": normalize_gt_type(row[index]),
                "坐标X": number(row[index + 1]),
                "坐标Y": number(row[index + 2]),
                "坐标Z": number(row[index + 3]),
                "尺寸数据": number(row[index + 4]),
            })
        name = str(row[0])
        output[name] = {
            "尺寸X": number(row[2]),
            "尺寸Y": number(row[3]),
            "尺寸Z": number(row[4]),
            "局部特征列表": features,
        }
        output[name].update(GT_DIMENSION_CORRECTIONS.get(name, {}))
    return output


def rmse(values: list[float]) -> float | None:
    return math.sqrt(sum(value * value for value in values) / len(values)) if values else None


def evaluate_prediction(name: str, prediction: dict, ground_truth: dict) -> dict:
    predicted = [
        feature for feature in prediction.get("局部特征列表", [])
        if feature.get("特征类型") in TARGET_TYPES
        and feature.get("作用") == ASSEMBLY_ROLE
    ]
    expected = [
        feature for feature in ground_truth.get("局部特征列表", [])
        if feature.get("特征类型") in TARGET_TYPES
    ]
    dimensions = [number(ground_truth.get(key)) or 0.0 for key in ("尺寸X", "尺寸Y", "尺寸Z")]
    diagonal = max(math.sqrt(sum(value * value for value in dimensions)), 1.0)
    candidates = []
    for pred_index, pred in enumerate(predicted):
        for gt_index, gt in enumerate(expected):
            if pred.get("特征类型") != gt.get("特征类型"):
                continue
            pred_coords = [number(pred.get(key)) for key in ("坐标X", "坐标Y", "坐标Z")]
            gt_coords = [number(gt.get(key)) for key in ("坐标X", "坐标Y", "坐标Z")]
            distance = (
                math.dist(pred_coords, gt_coords)
                if all(value is not None for value in pred_coords + gt_coords)
                else None
            )
            pred_size, gt_size = number(pred.get("尺寸数据")), number(gt.get("尺寸数据"))
            size_error = (
                abs(pred_size - gt_size)
                if pred_size is not None and gt_size is not None
                else None
            )
            normalized_distance = distance / diagonal if distance is not None else 0.0
            relative_size_error = size_error / max(abs(gt_size), 1.0) if size_error is not None else 0.0
            if normalized_distance <= 0.2 and relative_size_error <= 1.0:
                candidates.append((
                    normalized_distance + 0.25 * relative_size_error,
                    pred_index,
                    gt_index,
                    distance,
                    size_error,
                ))

    used_pred, used_gt, matches = set(), set(), []
    for _, pred_index, gt_index, distance, size_error in sorted(candidates):
        if pred_index in used_pred or gt_index in used_gt:
            continue
        used_pred.add(pred_index)
        used_gt.add(gt_index)
        matches.append((distance, size_error))

    gt_types = {feature.get("特征类型") for feature in expected}
    type_absent = sum(
        index not in used_pred and feature.get("特征类型") not in gt_types
        for index, feature in enumerate(predicted)
    )
    dimension_errors = [
        number(prediction.get(key)) - number(ground_truth.get(key))
        for key in ("尺寸X", "尺寸Y", "尺寸Z")
        if number(prediction.get(key)) is not None and number(ground_truth.get(key)) is not None
    ]
    return {
        "name": name,
        "difficulty": name.split("_", 1)[0],
        "predicted": len(predicted),
        "expected": len(expected),
        "matched": len(matches),
        "unmatched": len(predicted) - len(matches),
        "missed": len(expected) - len(matches),
        "type_absent": type_absent,
        "same_type_unmatched": len(predicted) - len(matches) - type_absent,
        "dimension_rmse": rmse(dimension_errors),
        "coordinate_rmse": rmse([value for value, _ in matches if value is not None]),
        "feature_size_rmse": rmse([value for _, value in matches if value is not None]),
    }


def summarize(rows: list[dict]) -> dict:
    predicted = sum(row["predicted"] for row in rows)
    expected = sum(row["expected"] for row in rows)
    matched = sum(row["matched"] for row in rows)

    def mean(key):
        values = [row[key] for row in rows if row[key] is not None]
        return sum(values) / len(values) if values else None

    return {
        "samples": len(rows),
        "predicted": predicted,
        "expected": expected,
        "matched": matched,
        "unmatched": predicted - matched,
        "missed": expected - matched,
        "type_absent": sum(row["type_absent"] for row in rows),
        "same_type_unmatched": sum(row["same_type_unmatched"] for row in rows),
        "precision": matched / predicted if predicted else None,
        "recall": matched / expected if expected else None,
        "unmatched_rate": (predicted - matched) / predicted if predicted else 0.0,
        "type_absent_rate": sum(row["type_absent"] for row in rows) / predicted if predicted else 0.0,
        "dimension_rmse": mean("dimension_rmse"),
        "coordinate_rmse": mean("coordinate_rmse"),
        "feature_size_rmse": mean("feature_size_rmse"),
    }


def evaluate_directory(results_dir: Path, ground_truth: dict, suffix: str) -> dict:
    rows = []
    for path in sorted(results_dir.glob(f"*{suffix}")):
        name = path.name[: -len(suffix)]
        if name not in ground_truth:
            continue
        rows.append(evaluate_prediction(name, json.loads(path.read_text(encoding="utf-8")), ground_truth[name]))
    grouped = defaultdict(list)
    for row in rows:
        grouped[row["difficulty"]].append(row)
    return {
        "overall": summarize(rows),
        "by_difficulty": {key: summarize(value) for key, value in sorted(grouped.items())},
        "per_sample": rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--results-dir", type=Path, required=True)
    parser.add_argument("--ground-truth", type=Path, required=True)
    parser.add_argument("--suffix", default="_refined_luna_visual_json_serial.txt")
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    report = evaluate_directory(
        args.results_dir.resolve(),
        load_ground_truth(args.ground_truth.resolve()),
        args.suffix,
    )
    rendered = json.dumps(report, ensure_ascii=False, indent=2)
    if args.output:
        args.output.write_text(rendered + "\n", encoding="utf-8")
    print(rendered)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
