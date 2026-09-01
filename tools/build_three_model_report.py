"""Build a reproducible Terra/Luna/GPT-5 Mini comparison workbook.

Terra and Luna are scored from their current raw JSON outputs with one shared
automatic matcher. GPT-5 Mini is kept in a separate legacy section because its
old workbook used manual labels and does not retain every false-positive feature
in the expanded feature columns.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
RESULT_ROOT = ROOT / "result"
GT_PATH = RESULT_ROOT / "old" / "exp1.xlsx"
OLD_CLASSIFICATION = RESULT_ROOT / "old" / "exp1_REMS_消融" / "准确率与幻觉率.xlsx"
OLD_COORDINATE = RESULT_ROOT / "old" / "exp1_REMS_消融" / "计算结果_RMSE_5行版.xlsx"
OLD_DIMENSION = RESULT_ROOT / "old" / "exp1_REMS_消融" / "计算结果_尺寸误差RMSE.xlsx"

MODES = [
    ("visual_only", "纯视觉"),
    ("visual_json_parallel", "视觉+JSON并行"),
    ("visual_json_serial", "视觉→JSON串行矫正"),
]
MODE_BY_OLD_LABEL = {
    "仅图片": "visual_only",
    "并行输出": "visual_json_parallel",
    "串行输出": "visual_json_serial",
}
OLD_SUFFIX_MODE = {
    "nojson": "visual_only",
    "integrated": "visual_json_parallel",
    "serial": "visual_json_serial",
}
MODE_NAME = dict(MODES)
MODEL_DIRS = {
    "Terra": ("terra", RESULT_ROOT / "terra_20260901"),
    "Luna": ("luna", RESULT_ROOT / "luna_20260901"),
}
TARGET_TYPES = {"孔", "柱", "槽", "倒角"}
ASSEMBLY_ROLE = "装配特征"
GT_DIMENSION_CORRECTIONS = {"hard_15": {"尺寸X": 32.42, "尺寸Y": 249.91}}
API_PRICES = {
    "Terra": {"input": 2.0, "output": 12.0},
    "Luna": {"input": 0.2, "output": 1.2},
}


def number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    return float(match.group()) if match else None


def difficulty(sample: str) -> str:
    return sample.split("_", 1)[0]


def normalize_gt_type(label) -> str | None:
    text = str(label or "")
    if "孔" in text:
        return "孔"
    if "槽" in text:
        return "槽"
    if "倒角" in text or "圆角" in text:
        return "倒角"
    if any(word in text for word in ("柱", "凸台", "外环", "主直径", "收细段", "半圆")):
        return "柱"
    return None


def load_ground_truth() -> dict[str, dict]:
    ws = load_workbook(GT_PATH, read_only=True, data_only=True)["Sheet2"]
    output = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        features = []
        for index in range(5, len(row), 5):
            if row[index] is None:
                continue
            features.append(
                {
                    "原始名称": str(row[index]),
                    "特征类型": normalize_gt_type(row[index]),
                    "坐标X": number(row[index + 1]),
                    "坐标Y": number(row[index + 2]),
                    "坐标Z": number(row[index + 3]),
                    "尺寸数据": number(row[index + 4]),
                    "作用": ASSEMBLY_ROLE,
                }
            )
        sample = str(row[0])
        output[sample] = {
            "名字": sample,
            "整体特征": row[1],
            "尺寸X": number(row[2]),
            "尺寸Y": number(row[3]),
            "尺寸Z": number(row[4]),
            "局部特征列表": features,
        }
        output[sample].update(GT_DIMENSION_CORRECTIONS.get(sample, {}))
    return output


def load_predictions(model: str, tag: str, directory: Path) -> list[dict]:
    rows = []
    for mode_key, mode_name in MODES:
        suffix = f"_refined_{tag}_{mode_key}.txt"
        for path in sorted(directory.glob(f"*{suffix}")):
            data = json.loads(path.read_text(encoding="utf-8"))
            data.update(
                {
                    "模型": model,
                    "名字": path.name[: -len(suffix)],
                    "流程键": mode_key,
                    "流程": mode_name,
                    "结果文件": path.name,
                }
            )
            rows.append(data)
    return rows


def load_runtime(model: str, directory: Path) -> list[dict]:
    rows = []
    for mode_key, mode_name in MODES:
        with (directory / f"metrics_{mode_key}.csv").open(encoding="utf-8-sig", newline="") as handle:
            for source in csv.DictReader(handle):
                rows.append(
                    {
                        "模型": model,
                        "文件名": source["base_name"],
                        "难度": difficulty(source["base_name"]),
                        "流程键": mode_key,
                        "流程": mode_name,
                        "模型耗时(s)": number(source["codex_time_s"]) or 0.0,
                        "输入tokens": int(number(source["prompt_tokens"]) or 0),
                        "输出tokens": int(number(source["completion_tokens"]) or 0),
                        "总tokens": int(number(source["total_tokens"]) or 0),
                        "状态": source["status"],
                        "错误": source["error"],
                    }
                )
    return rows


def rmse(values: list[float]) -> float | None:
    return math.sqrt(sum(value * value for value in values) / len(values)) if values else None


def overall_dimension_rmse(pred: dict, gt: dict) -> float | None:
    errors = []
    for key in ("尺寸X", "尺寸Y", "尺寸Z"):
        pred_value, gt_value = number(pred.get(key)), number(gt.get(key))
        if pred_value is not None and gt_value is not None:
            errors.append(pred_value - gt_value)
    return rmse(errors)


def evaluate_prediction(pred: dict, gt: dict) -> dict:
    pred_features = [
        feature
        for feature in pred.get("局部特征列表", [])
        if feature.get("特征类型") in TARGET_TYPES and feature.get("作用") == ASSEMBLY_ROLE
    ]
    gt_features = [
        feature
        for feature in gt.get("局部特征列表", [])
        if feature.get("特征类型") in TARGET_TYPES
    ]
    dims = [number(gt.get(key)) or 0.0 for key in ("尺寸X", "尺寸Y", "尺寸Z")]
    diagonal = max(math.sqrt(sum(value * value for value in dims)), 1.0)
    candidates = []
    for pred_index, pred_feature in enumerate(pred_features):
        for gt_index, gt_feature in enumerate(gt_features):
            if pred_feature.get("特征类型") != gt_feature.get("特征类型"):
                continue
            pred_coords = [number(pred_feature.get(key)) for key in ("坐标X", "坐标Y", "坐标Z")]
            gt_coords = [number(gt_feature.get(key)) for key in ("坐标X", "坐标Y", "坐标Z")]
            if all(value is not None for value in pred_coords + gt_coords):
                distance = math.sqrt(sum((left - right) ** 2 for left, right in zip(pred_coords, gt_coords)))
                normalized_distance = distance / diagonal
            else:
                distance = None
                normalized_distance = 0.0
            pred_size = number(pred_feature.get("尺寸数据"))
            gt_size = number(gt_feature.get("尺寸数据"))
            if pred_size is not None and gt_size is not None:
                size_error = abs(pred_size - gt_size)
                relative_size_error = size_error / max(abs(gt_size), 1.0)
            else:
                size_error = None
                relative_size_error = 0.0
            if normalized_distance <= 0.20 and relative_size_error <= 1.0:
                candidates.append(
                    (
                        normalized_distance + 0.25 * relative_size_error,
                        pred_index,
                        gt_index,
                        distance,
                        size_error,
                    )
                )

    used_pred, used_gt, matches = set(), set(), []
    for _, pred_index, gt_index, distance, size_error in sorted(candidates):
        if pred_index in used_pred or gt_index in used_gt:
            continue
        used_pred.add(pred_index)
        used_gt.add(gt_index)
        matches.append((pred_index, gt_index, distance, size_error))

    gt_types = {feature.get("特征类型") for feature in gt_features}
    type_absent = sum(
        index not in used_pred and feature.get("特征类型") not in gt_types
        for index, feature in enumerate(pred_features)
    )
    same_type_unmatched = len(pred_features) - len(matches) - type_absent
    tp, fp, fn = len(matches), len(pred_features) - len(matches), len(gt_features) - len(matches)
    return {
        "模型": pred["模型"],
        "文件名": pred["名字"],
        "难度": difficulty(pred["名字"]),
        "流程键": pred["流程键"],
        "流程": pred["流程"],
        "预测装配特征数": len(pred_features),
        "GT装配特征数": len(gt_features),
        "匹配数": tp,
        "未匹配预测数": fp,
        "漏检数": fn,
        "类型不存在数": type_absent,
        "同类型未匹配数": same_type_unmatched,
        "精确率": tp / (tp + fp) if tp + fp else None,
        "召回率": tp / (tp + fn) if tp + fn else None,
        "未匹配率": fp / (tp + fp) if tp + fp else 0.0,
        "类型幻觉下界": type_absent / (tp + fp) if tp + fp else 0.0,
        "整体尺寸RMSE": overall_dimension_rmse(pred, gt),
        "坐标RMSE": rmse([row[2] for row in matches if row[2] is not None]),
        "特征尺寸RMSE": rmse([row[3] for row in matches if row[3] is not None]),
    }


def mean_available(rows: list[dict], key: str) -> float | None:
    values = [row[key] for row in rows if row.get(key) is not None]
    return sum(values) / len(values) if values else None


def summarize(rows: list[dict]) -> dict:
    tp = sum(row["匹配数"] for row in rows)
    fp = sum(row["未匹配预测数"] for row in rows)
    fn = sum(row["漏检数"] for row in rows)
    type_absent = sum(row["类型不存在数"] for row in rows)
    same_type = sum(row["同类型未匹配数"] for row in rows)
    return {
        "样本数": len(rows),
        "预测装配特征数": tp + fp,
        "GT装配特征数": tp + fn,
        "匹配数": tp,
        "未匹配预测数": fp,
        "漏检数": fn,
        "类型不存在数": type_absent,
        "同类型未匹配数": same_type,
        "精确率": tp / (tp + fp) if tp + fp else None,
        "召回率": tp / (tp + fn) if tp + fn else None,
        "未匹配率": fp / (tp + fp) if tp + fp else 0.0,
        "类型幻觉下界": type_absent / (tp + fp) if tp + fp else 0.0,
        "同类型未匹配率": same_type / (tp + fp) if tp + fp else 0.0,
        "整体尺寸RMSE": mean_available(rows, "整体尺寸RMSE"),
        "坐标RMSE": mean_available(rows, "坐标RMSE"),
        "特征尺寸RMSE": mean_available(rows, "特征尺寸RMSE"),
    }


def load_old_classification() -> tuple[list[dict], list[dict]]:
    workbook = load_workbook(OLD_CLASSIFICATION, read_only=True, data_only=True)
    detail_sheet = workbook["Sheet1"]
    details = []
    pattern = re.compile(r"(easy|medium|hard)_(\d+)_refined_(integrated|nojson|serial)\.txt")
    for row in detail_sheet.iter_rows(min_row=2, values_only=True):
        match = pattern.fullmatch(str(row[0] or ""))
        if not match:
            continue
        mode_key = OLD_SUFFIX_MODE[match.group(3)]
        details.append(
            {
                "文件名": f"{match.group(1)}_{int(match.group(2))}",
                "难度": match.group(1),
                "流程键": mode_key,
                "流程": MODE_NAME[mode_key],
                "总特征": int(number(row[1]) or 0),
                "识别": int(number(row[2]) or 0),
                "其他": int(number(row[3]) or 0),
                "不存在": int(number(row[4]) or 0),
                "旧单件幻觉率": number(row[6]),
                "旧单件有效输出率": number(row[7]),
                "旧单件召回率": number(row[8]),
            }
        )

    summary_sheet = workbook["Sheet2"]
    published = {mode_key: {"流程": MODE_NAME[mode_key]} for mode_key, _ in MODES}
    for column in range(7, 10):
        published[MODE_BY_OLD_LABEL[summary_sheet.cell(2, column).value]]["旧展示幻觉率"] = number(
            summary_sheet.cell(3, column).value
        )
        published[MODE_BY_OLD_LABEL[summary_sheet.cell(7, column).value]]["旧展示有效输出率"] = number(
            summary_sheet.cell(8, column).value
        )
        published[MODE_BY_OLD_LABEL[summary_sheet.cell(11, column).value]]["旧展示召回率"] = number(
            summary_sheet.cell(12, column).value
        )

    summaries = []
    for mode_key, mode_name in MODES:
        rows = [row for row in details if row["流程键"] == mode_key]
        total = sum(row["总特征"] for row in rows)
        nonexistent = sum(row["不存在"] for row in rows)
        valid_values = [row["旧单件有效输出率"] for row in rows if row["旧单件有效输出率"] is not None]
        recall_values = [row["旧单件召回率"] for row in rows if row["旧单件召回率"] is not None]
        summaries.append(
            {
                "流程键": mode_key,
                "流程": mode_name,
                "样本数": len(rows),
                "总特征": total,
                "识别": sum(row["识别"] for row in rows),
                "其他": sum(row["其他"] for row in rows),
                "不存在": nonexistent,
                "按原始计数幻觉率": nonexistent / total if total else None,
                "按旧单件宏平均有效输出率": sum(valid_values) / len(valid_values) if valid_values else None,
                "按旧单件宏平均召回率": sum(recall_values) / len(recall_values) if recall_values else None,
                **published[mode_key],
            }
        )
    return details, summaries


def load_old_rmse() -> list[dict]:
    output = {(level, mode_key): {"难度": level, "流程键": mode_key, "流程": MODE_NAME[mode_key]} for level in ("easy", "medium", "hard") for mode_key, _ in MODES}
    coordinate = load_workbook(OLD_COORDINATE, read_only=True, data_only=True)["Sheet2"]
    levels = [coordinate.cell(46, column).value for column in range(14, 17)]
    for row in range(47, 50):
        mode_key = MODE_BY_OLD_LABEL[coordinate.cell(row, 13).value]
        for column, level in zip(range(14, 17), levels):
            output[(level, mode_key)]["坐标RMSE"] = number(coordinate.cell(row, column).value)
    dimension = load_workbook(OLD_DIMENSION, read_only=True, data_only=True)["Sheet1"]
    levels = [dimension.cell(37, column).value for column in range(8, 11)]
    for row in range(38, 41):
        mode_key = MODE_BY_OLD_LABEL[dimension.cell(row, 7).value]
        for column, level in zip(range(8, 11), levels):
            output[(level, mode_key)]["整体尺寸RMSE"] = number(dimension.cell(row, column).value)
    levels = [dimension.cell(47, column).value for column in range(8, 11)]
    for row in range(48, 51):
        mode_key = MODE_BY_OLD_LABEL[dimension.cell(row, 7).value]
        for column, level in zip(range(8, 11), levels):
            output[(level, mode_key)]["特征尺寸RMSE"] = number(dimension.cell(row, column).value)
    return list(output.values())


def style_sheet(sheet, freeze="A2") -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = freeze
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, min(sheet.max_row, 120) + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max(map(len, values), default=8) + 2, 10), 36)


def append_table(sheet, headers: list[str], rows: list[dict]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    style_sheet(sheet)


def add_four_row_sheet(workbook: Workbook, title: str, model: str, predictions: list[dict], ground_truth: dict[str, dict], evaluations: list[dict]) -> None:
    sheet = workbook.create_sheet(title)
    headers = [
        "文件名", "行类型", "流程", "模型", "整体特征", "尺寸X", "尺寸Y", "尺寸Z",
        "局部特征总数", "预测装配特征数", "GT装配特征数", "匹配数", "未匹配预测数",
        "漏检数", "类型不存在数", "同类型未匹配数", "精确率", "召回率", "未匹配率",
        "类型幻觉下界", "整体尺寸RMSE", "坐标RMSE", "特征尺寸RMSE", "局部特征JSON", "结果文件",
    ]
    sheet.append(headers)
    pred_by_key = {(row["名字"], row["流程键"]): row for row in predictions if row["模型"] == model}
    eval_by_key = {(row["文件名"], row["流程键"]): row for row in evaluations if row["模型"] == model}
    names = sorted(ground_truth, key=lambda value: (difficulty(value), int(value.split("_")[1])))
    for sample in names:
        for mode_key, mode_name in MODES:
            pred = pred_by_key[(sample, mode_key)]
            evaluation = eval_by_key[(sample, mode_key)]
            sheet.append(
                [
                    sample, "预测", mode_name, model, pred.get("整体特征"), pred.get("尺寸X"), pred.get("尺寸Y"), pred.get("尺寸Z"),
                    len(pred.get("局部特征列表", [])),
                    *[evaluation[key] for key in headers[9:23]],
                    json.dumps(pred.get("局部特征列表", []), ensure_ascii=False), pred["结果文件"],
                ]
            )
        gt = ground_truth[sample]
        sheet.append(
            [
                sample, "参考答案", "参考答案", "GT", gt.get("整体特征"), gt.get("尺寸X"), gt.get("尺寸Y"), gt.get("尺寸Z"),
                len(gt.get("局部特征列表", [])), None, len(gt.get("局部特征列表", [])),
                None, None, None, None, None, None, None, None, None, None, None, None,
                json.dumps(gt.get("局部特征列表", []), ensure_ascii=False), GT_PATH.name,
            ]
        )
    style_sheet(sheet)
    group_border = Border(bottom=Side(style="medium", color="5B9BD5"))
    for row_number in range(5, sheet.max_row + 1, 4):
        for cell in sheet[row_number]:
            cell.border = group_border
    for row in sheet.iter_rows(min_row=2, min_col=17, max_col=20):
        for cell in row:
            cell.number_format = "0.00%"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        type=Path,
        default=RESULT_ROOT / "Terra_Luna_GPT5Mini_三模型分析整合.xlsx",
    )
    args = parser.parse_args()

    ground_truth = load_ground_truth()
    predictions, runtime_rows = [], []
    for model, (tag, directory) in MODEL_DIRS.items():
        predictions.extend(load_predictions(model, tag, directory))
        runtime_rows.extend(load_runtime(model, directory))
    if len(ground_truth) != 29 or len(predictions) != 174 or len(runtime_rows) != 174:
        raise RuntimeError(
            f"输入数量异常: GT={len(ground_truth)}, predictions={len(predictions)}, runtime={len(runtime_rows)}"
        )
    if any(row["状态"] != "OK" for row in runtime_rows):
        raise RuntimeError("Terra/Luna 运行指标中存在非 OK 记录")

    evaluations = [evaluate_prediction(pred, ground_truth[pred["名字"]]) for pred in predictions]
    grouped_evaluations = defaultdict(list)
    grouped_runtime = defaultdict(list)
    for row in evaluations:
        grouped_evaluations[(row["模型"], row["流程键"])].append(row)
    for row in runtime_rows:
        grouped_runtime[(row["模型"], row["流程键"])].append(row)

    summaries = []
    for model in MODEL_DIRS:
        for mode_key, mode_name in MODES:
            summary = summarize(grouped_evaluations[(model, mode_key)])
            runs = grouped_runtime[(model, mode_key)]
            input_tokens = sum(row["输入tokens"] for row in runs)
            output_tokens = sum(row["输出tokens"] for row in runs)
            estimate = input_tokens / 1_000_000 * API_PRICES[model]["input"] + output_tokens / 1_000_000 * API_PRICES[model]["output"]
            summaries.append(
                {
                    "模型": model,
                    "流程键": mode_key,
                    "流程": mode_name,
                    **summary,
                    "模型总耗时(s)": sum(row["模型耗时(s)"] for row in runs),
                    "模型平均耗时(s)": sum(row["模型耗时(s)"] for row in runs) / len(runs),
                    "输入tokens": input_tokens,
                    "输出tokens": output_tokens,
                    "总tokens": sum(row["总tokens"] for row in runs),
                    "按API标准价估算($)": estimate,
                }
            )

    difficulty_rows = []
    for model in MODEL_DIRS:
        for level in ("easy", "medium", "hard"):
            for mode_key, mode_name in MODES:
                rows = [row for row in grouped_evaluations[(model, mode_key)] if row["难度"] == level]
                difficulty_rows.append({"模型": model, "难度": level, "流程": mode_name, **summarize(rows)})

    old_details, old_summaries = load_old_classification()
    old_rmse = load_old_rmse()

    workbook = Workbook()
    workbook.remove(workbook.active)

    sheet = workbook.create_sheet("结论摘要")
    sheet.append(["项目", "结论"])
    summary_by_key = {(row["模型"], row["流程键"]): row for row in summaries}
    terra_serial = summary_by_key[("Terra", "visual_json_serial")]
    luna_parallel = summary_by_key[("Luna", "visual_json_parallel")]
    terra_parallel = summary_by_key[("Terra", "visual_json_parallel")]
    notes = [
        ("实验完成", "Terra 与 Luna 均为 29 个 STL × 3 条流程，共 174 份结构化结果；所有运行状态为 OK。串行流程每件两轮模型调用。"),
        ("首选：低误报/均衡", f"Terra 串行：精确率 {terra_serial['精确率']:.2%}、召回率 {terra_serial['召回率']:.2%}、未匹配率 {terra_serial['未匹配率']:.2%}、类型幻觉下界 {terra_serial['类型幻觉下界']:.2%}。"),
        ("首选：高召回", f"Luna 并行：召回率 {luna_parallel['召回率']:.2%} 为六组最高，但精确率仅 {luna_parallel['精确率']:.2%}、未匹配率 {luna_parallel['未匹配率']:.2%}，适合宁可多报、后续有人或规则复核的场景。"),
        ("首选：几何尺寸", f"并行流程能直接利用几何 JSON。Terra 并行的整体尺寸 RMSE={terra_parallel['整体尺寸RMSE']:.3f}、特征尺寸 RMSE={terra_parallel['特征尺寸RMSE']:.3f}；Luna 并行坐标 RMSE={luna_parallel['坐标RMSE']:.3f}。"),
        ("不建议直接称为幻觉率", "自动评分的“未匹配率”混合了真实不存在、同类型但定位/尺寸超阈值、重复预测。报告另列“类型幻觉下界”；严格幻觉率仍需人工逐项确认。"),
        ("5 mini 比较限制", "5 mini 是旧批次：分类表含28件、与当前实验重合27件；RMSE源表含29件、与当前实验重合28件。人工分类与自动评分不同，展开明细还缺少误检特征。因此只保留旧发布结果和原始计数，不能与 Terra/Luna 做严格差值或显著性结论。"),
        ("官方定位", "OpenAI Docs：Terra 是平衡智能与成本的 5.6 mini 档；Luna 是低成本高吞吐的 5.6 nano 档；旧 GPT-5 Mini 适合定义清楚、提示精确的任务，官方建议新低延迟/高吞吐任务从 Terra 开始评估。"),
        ("部署建议", "若必须自动直出装配特征，优先 Terra 串行；若有二阶段校验器且漏检代价最高，可选 Luna 并行；纯视觉仅适合作为无几何 JSON 时的降级路径。"),
    ]
    for note in notes:
        sheet.append(note)
    style_sheet(sheet)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 110

    summary_headers = [
        "模型", "流程", "样本数", "预测装配特征数", "GT装配特征数", "匹配数", "未匹配预测数", "漏检数",
        "类型不存在数", "同类型未匹配数", "精确率", "召回率", "未匹配率", "类型幻觉下界", "同类型未匹配率",
        "整体尺寸RMSE", "坐标RMSE", "特征尺寸RMSE", "模型总耗时(s)", "模型平均耗时(s)",
        "输入tokens", "输出tokens", "总tokens", "按API标准价估算($)",
    ]
    sheet = workbook.create_sheet("Terra_Luna同口径")
    append_table(sheet, summary_headers, summaries)
    for row in sheet.iter_rows(min_row=2, min_col=11, max_col=15):
        for cell in row:
            cell.number_format = "0.00%"

    difficulty_headers = [
        "模型", "难度", "流程", "样本数", "预测装配特征数", "GT装配特征数", "匹配数", "未匹配预测数",
        "漏检数", "类型不存在数", "同类型未匹配数", "精确率", "召回率", "未匹配率", "类型幻觉下界",
        "同类型未匹配率", "整体尺寸RMSE", "坐标RMSE", "特征尺寸RMSE",
    ]
    sheet = workbook.create_sheet("Terra_Luna难度分层")
    append_table(sheet, difficulty_headers, difficulty_rows)
    for row in sheet.iter_rows(min_row=2, min_col=12, max_col=16):
        for cell in row:
            cell.number_format = "0.00%"

    evaluation_headers = [
        "模型", "文件名", "难度", "流程", "预测装配特征数", "GT装配特征数", "匹配数", "未匹配预测数",
        "漏检数", "类型不存在数", "同类型未匹配数", "精确率", "召回率", "未匹配率", "类型幻觉下界",
        "整体尺寸RMSE", "坐标RMSE", "特征尺寸RMSE",
    ]
    sheet = workbook.create_sheet("Terra_Luna逐件评分")
    append_table(sheet, evaluation_headers, evaluations)
    for row in sheet.iter_rows(min_row=2, min_col=12, max_col=15):
        for cell in row:
            cell.number_format = "0.00%"

    runtime_headers = ["模型", "文件名", "难度", "流程", "模型耗时(s)", "输入tokens", "输出tokens", "总tokens", "状态", "错误"]
    sheet = workbook.create_sheet("Terra_Luna运行明细")
    append_table(sheet, runtime_headers, runtime_rows)

    add_four_row_sheet(workbook, "Terra四行明细", "Terra", predictions, ground_truth, evaluations)
    add_four_row_sheet(workbook, "Luna四行明细", "Luna", predictions, ground_truth, evaluations)

    old_summary_headers = [
        "流程", "样本数", "总特征", "识别", "其他", "不存在", "按原始计数幻觉率",
        "按旧单件宏平均有效输出率", "按旧单件宏平均召回率", "旧展示幻觉率", "旧展示有效输出率", "旧展示召回率",
    ]
    sheet = workbook.create_sheet("5mini旧分类汇总")
    append_table(sheet, old_summary_headers, old_summaries)
    for row in sheet.iter_rows(min_row=2, min_col=7, max_col=12):
        for cell in row:
            cell.number_format = "0.00%"

    old_detail_headers = ["文件名", "难度", "流程", "总特征", "识别", "其他", "不存在", "旧单件幻觉率", "旧单件有效输出率", "旧单件召回率"]
    sheet = workbook.create_sheet("5mini旧分类逐件")
    append_table(sheet, old_detail_headers, old_details)
    for row in sheet.iter_rows(min_row=2, min_col=8, max_col=10):
        for cell in row:
            cell.number_format = "0.00%"

    old_rmse_headers = ["难度", "流程", "整体尺寸RMSE", "坐标RMSE", "特征尺寸RMSE"]
    sheet = workbook.create_sheet("5mini旧RMSE")
    append_table(sheet, old_rmse_headers, old_rmse)

    sheet = workbook.create_sheet("三模型并列_谨慎解读")
    side_headers = ["模型", "流程", "评价来源", "样本", "精确率或有效输出率", "召回率", "未匹配率或旧幻觉率", "可比性说明"]
    side_rows = []
    for row in summaries:
        side_rows.append(
            {
                "模型": row["模型"], "流程": row["流程"], "评价来源": "统一自动匹配",
                "样本": "当前29件", "精确率或有效输出率": row["精确率"], "召回率": row["召回率"],
                "未匹配率或旧幻觉率": row["未匹配率"], "可比性说明": "Terra/Luna之间可直接比较；未匹配率不等于严格幻觉率。",
            }
        )
    for row in old_summaries:
        side_rows.append(
            {
                "模型": "GPT-5 Mini", "流程": row["流程"], "评价来源": "旧人工分类/旧展示",
                "样本": "旧分类28件（与当前重合27件）", "精确率或有效输出率": row["旧展示有效输出率"],
                "召回率": row["旧展示召回率"], "未匹配率或旧幻觉率": row["旧展示幻觉率"],
                "可比性说明": "分类来源、提示/流水线、样本均不完全相同，只能作历史参照。",
            }
        )
    append_table(sheet, side_headers, side_rows)
    for row in sheet.iter_rows(min_row=2, min_col=5, max_col=7):
        for cell in row:
            cell.number_format = "0.00%"

    sheet = workbook.create_sheet("指标与来源说明")
    sheet.append(["项目", "定义/来源"])
    explanations = [
        ("当前实验", "Terra/Luna：同一29件 STL、同一三条 pipeline、medium reasoning、同一 GT、同一自动匹配器。"),
        ("精确率", "自动匹配成功数 / 预测装配特征数。"),
        ("召回率", "自动匹配成功数 / GT 装配特征数。"),
        ("未匹配率", "未匹配预测装配特征数 / 预测装配特征数；包含严格幻觉、重复预测、类型/定位/尺寸错误，不应直接命名为严格幻觉率。"),
        ("类型幻觉下界", "未匹配预测中，其特征类型在该件 GT 完全不存在的数量 / 预测装配特征数；仍是自动下界，不替代人工确认。"),
        ("同类型未匹配", "GT 存在相同类型，但因数量、坐标或尺寸未通过一对一匹配的预测。"),
        ("匹配门槛", "同类型；坐标距离≤GT包络对角线20%；尺寸相对误差≤100%；候选按归一化距离+0.25×尺寸相对误差贪心一对一匹配。"),
        ("RMSE汇总", "先按单件计算，再对有值的单件做算术平均，与旧 RMSE 报表的分层汇总方式一致。"),
        ("API价估算", "仅按当前 OpenAI API 标准输入/输出单价和记录 token 粗估，未扣除缓存折扣；本次实际通过 Codex ChatGPT/OAuth 运行，不代表实际账单。"),
        ("Terra官方", "https://developers.openai.com/api/docs/models/gpt-5.6-terra；平衡智能与成本，约对应早期 mini 档；$2/M input、$12/M output。"),
        ("Luna官方", "https://developers.openai.com/api/docs/models/gpt-5.6-luna；低成本高吞吐，约对应早期 nano 档；$0.20/M input、$1.20/M output。"),
        ("GPT-5 Mini官方", "https://developers.openai.com/api/docs/models/gpt-5-mini；适合定义清楚、提示精确的任务；官方建议多数新低延迟/高吞吐任务从 Terra 开始。"),
        ("5 mini旧分类", "读取 result/old/exp1_REMS_消融/准确率与幻觉率.xlsx；识别/其他/不存在为人工分类。旧‘精确率’实际是有效输出率=(识别+其他)/总特征。"),
        ("5 mini旧RMSE", "读取旧坐标和尺寸 RMSE 工作簿的发布汇总区；不重新套用当前自动匹配器。"),
        ("样本差异", "旧5 mini分类表含28件：有easy_12、无easy_8和hard_32，与当前重合27件。旧RMSE/三流程源表另含hard_32，共29件，与当前重合28件。"),
        ("GT修正", "hard_15 原GT整体尺寸X/Y写反；当前统一按X=32.42、Y=249.91、Z=34计算。"),
    ]
    for explanation in explanations:
        sheet.append(explanation)
    style_sheet(sheet)
    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 115

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(args.output.resolve())


if __name__ == "__main__":
    main()
