"""Evaluate consensus outputs with the project's shared 29-part matcher."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

import build_sample_report as benchmark


BASELINE_MODES = (
    "visual_only",
    "visual_json_parallel",
    "visual_json_serial",
)


GT_DIMENSION_CORRECTIONS = {"hard_15": {"尺寸X": 32.42, "尺寸Y": 249.91}}


def f1(summary: dict) -> float | None:
    precision = summary.get("精确率")
    recall = summary.get("召回率")
    if precision is None or recall is None or precision + recall == 0:
        return None
    return 2 * precision * recall / (precision + recall)


def load_prediction(path: Path, sample: str, mode: str, flow_name: str) -> dict:
    prediction = json.loads(path.read_text(encoding="utf-8"))
    prediction.update(
        {
            "模型": "Terra",
            "名字": sample,
            "流程键": mode,
            "流程": flow_name,
        }
    )
    return prediction


def sample_names(ground_truth_path: Path) -> list[str]:
    worksheet = load_workbook(ground_truth_path, read_only=True, data_only=True)["Sheet2"]
    return [str(row[0]) for row in worksheet.iter_rows(min_row=2, values_only=True) if row[0]]


def mean_available(rows: list[dict], key: str) -> float | None:
    values = [row[key] for row in rows if row.get(key) is not None]
    return sum(values) / len(values) if values else None


def summarize(rows: list[dict]) -> dict:
    matches = sum(row["自动匹配数"] for row in rows)
    false_positives = sum(row["未匹配预测数"] for row in rows)
    false_negatives = sum(row["漏检数"] for row in rows)
    summary = {
        "样本数": len(rows),
        "预测装配特征数": matches + false_positives,
        "GT装配特征数": matches + false_negatives,
        "匹配数": matches,
        "未匹配预测数": false_positives,
        "漏检数": false_negatives,
        "精确率": matches / (matches + false_positives) if matches + false_positives else None,
        "召回率": matches / (matches + false_negatives) if matches + false_negatives else None,
        "整体尺寸RMSE": mean_available(rows, "整体尺寸RMSE"),
        "坐标RMSE": mean_available(rows, "坐标RMSE"),
        "特征尺寸RMSE": mean_available(rows, "特征尺寸RMSE"),
    }
    summary["F1"] = f1(summary)
    return summary


def evaluate_directory(
    baseline_dir: Path,
    consensus_dir: Path,
    ground_truth_path: Path,
    model_tag: str,
    profile: str,
) -> dict:
    rows_by_flow: dict[str, list[dict]] = {mode: [] for mode in BASELINE_MODES}
    consensus_key = f"consensus_{profile}"
    rows_by_flow[consensus_key] = []

    for sample in sample_names(ground_truth_path):
        truth = benchmark.load_ground_truth(ground_truth_path, sample)
        truth.update(GT_DIMENSION_CORRECTIONS.get(sample, {}))
        paths = {
            mode: baseline_dir / f"{sample}_refined_{model_tag}_{mode}.txt"
            for mode in BASELINE_MODES
        }
        paths[consensus_key] = consensus_dir / f"{sample}_refined_{model_tag}_{consensus_key}.txt"
        if not all(path.is_file() for path in paths.values()):
            continue

        for mode, path in paths.items():
            flow_name = mode if mode in BASELINE_MODES else f"共识融合-{profile}"
            prediction = load_prediction(path, sample, mode, flow_name)
            metrics = benchmark.evaluate(prediction, truth)
            metrics["难度"] = sample.split("_", 1)[0]
            rows_by_flow[mode].append(metrics)

    output = {
        "model_tag": model_tag,
        "profile": profile,
        "ground_truth": str(ground_truth_path),
        "samples": len(rows_by_flow[consensus_key]),
        "flows": {},
    }
    for mode, rows in rows_by_flow.items():
        summary = summarize(rows)
        difficulty_summaries = {}
        for level in ("easy", "medium", "hard"):
            level_rows = [row for row in rows if row["难度"] == level]
            level_summary = summarize(level_rows)
            difficulty_summaries[level] = level_summary
        output["flows"][mode] = {
            "summary": summary,
            "by_difficulty": difficulty_summaries,
        }
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--baseline-dir", type=Path, required=True)
    parser.add_argument("--consensus-dir", type=Path, required=True)
    parser.add_argument("--ground-truth", type=Path, required=True)
    parser.add_argument("--model-tag", default="terra")
    parser.add_argument("--profile", choices=("balanced", "precision"), default="balanced")
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    report = evaluate_directory(
        args.baseline_dir.resolve(),
        args.consensus_dir.resolve(),
        args.ground_truth.resolve(),
        args.model_tag,
        args.profile,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(args.output.resolve())


if __name__ == "__main__":
    main()
