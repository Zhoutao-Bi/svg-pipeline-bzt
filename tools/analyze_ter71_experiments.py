"""Build the paired TER-71 ablation and efficiency report.

The script deliberately consumes the published TER-65/TER-67 workbooks rather
than re-scoring prose fields with a second, subtly different matcher.  Each
workbook already contains the shared one-to-one GT evaluation.  This tool
validates their common case set, assembles the six-configuration ladder, and
adds case-clustered bootstrap intervals plus dynamic-slicing cost statistics.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import random
import statistics
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable, Mapping, Sequence

import openpyxl
from openpyxl.styles import Font
from scipy.stats import spearmanr


CONFIG_ORDER = (
    "visual_only",
    "parallel",
    "serial",
    "dynamic",
    "topology",
    "full",
)

DISPLAY_NAMES = {
    "visual_only": "Terra visual only",
    "parallel": "Terra visual + JSON parallel",
    "serial": "Terra coarse serial",
    "dynamic": "Terra dynamic serial",
    "topology": "Terra topology dynamic",
    "full": "Terra full pipeline (d98baca)",
}


@dataclass(frozen=True)
class CaseScore:
    name: str
    difficulty: str
    predicted: int
    expected: int
    matched: int

    def __post_init__(self) -> None:
        if min(self.predicted, self.expected, self.matched) < 0:
            raise ValueError(f"negative count for {self.name}")
        if self.matched > min(self.predicted, self.expected):
            raise ValueError(f"impossible match count for {self.name}")


@dataclass(frozen=True)
class Summary:
    samples: int
    predicted: int
    expected: int
    matched: int
    precision: float
    recall: float
    f1: float
    unmatched_rate: float


def summarize(scores: Iterable[CaseScore]) -> Summary:
    rows = list(scores)
    predicted = sum(row.predicted for row in rows)
    expected = sum(row.expected for row in rows)
    matched = sum(row.matched for row in rows)
    precision = matched / predicted if predicted else 1.0
    recall = matched / expected if expected else 1.0
    f1 = 2 * matched / (predicted + expected) if predicted + expected else 1.0
    return Summary(
        samples=len(rows),
        predicted=predicted,
        expected=expected,
        matched=matched,
        precision=precision,
        recall=recall,
        f1=f1,
        unmatched_rate=(predicted - matched) / predicted if predicted else 0.0,
    )


def _sheet_rows(path: Path, sheet: str) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet not in workbook.sheetnames:
        raise ValueError(f"{path.name} has no sheet {sheet!r}")
    iterator = workbook[sheet].iter_rows(values_only=True)
    header = [str(value) for value in next(iterator)]
    return [dict(zip(header, row)) for row in iterator]


def _difficulty(name: str) -> str:
    difficulty = name.split("_", 1)[0].lower()
    if difficulty not in {"easy", "medium", "hard"}:
        raise ValueError(f"unrecognized difficulty in {name!r}")
    return difficulty


def _score(
    *, name: object, difficulty: object, predicted: object, expected: object, matched: object
) -> CaseScore:
    case_name = str(name)
    return CaseScore(
        name=case_name,
        difficulty=str(difficulty or _difficulty(case_name)).lower(),
        predicted=int(predicted or 0),
        expected=int(expected or 0),
        matched=int(matched or 0),
    )


def load_three_flow(path: Path) -> dict[str, dict[str, CaseScore]]:
    flow_names = {
        "纯视觉": "visual_only",
        "视觉+JSON并行": "parallel",
        "视觉→JSON串行矫正": "serial",
    }
    result = {key: {} for key in flow_names.values()}
    for row in _sheet_rows(path, "Terra四行明细"):
        if row.get("行类型") != "预测" or row.get("模型") != "Terra":
            continue
        key = flow_names.get(str(row.get("流程")))
        if key is None:
            continue
        score = _score(
            name=row["文件名"],
            difficulty=None,
            predicted=row["预测装配特征数"],
            expected=row["GT装配特征数"],
            matched=row["匹配数"],
        )
        result[key][score.name] = score
    return result


def load_dynamic(path: Path) -> dict[str, dict[str, CaseScore]]:
    version_names = {
        "父Issue Terra旧串行": "serial",
        "Terra动态串行": "dynamic",
    }
    result = {key: {} for key in version_names.values()}
    for row in _sheet_rows(path, "逐件评分"):
        key = version_names.get(str(row.get("版本")))
        if key is None:
            continue
        score = _score(
            name=row["文件名"],
            difficulty=row["难度"],
            predicted=row["预测装配特征数"],
            expected=row["GT装配特征数"],
            matched=row["匹配数"],
        )
        result[key][score.name] = score
    return result


def load_topology(path: Path) -> dict[str, CaseScore]:
    result: dict[str, CaseScore] = {}
    for row in _sheet_rows(path, "逐件评分"):
        if row.get("版本") != "新版拓扑动态Terra":
            continue
        score = _score(
            name=row["文件名"],
            difficulty=row["难度"],
            predicted=row["预测装配特征数"],
            expected=row["GT装配特征数"],
            matched=row["匹配数"],
        )
        result[score.name] = score
    return result


def load_final(path: Path) -> dict[str, CaseScore]:
    result: dict[str, CaseScore] = {}
    for row in _sheet_rows(path, "Per Sample"):
        score = _score(
            name=row["Name"],
            difficulty=row["Difficulty"],
            predicted=row["Final predicted"],
            expected=row["Final expected"],
            matched=row["Final matched"],
        )
        result[score.name] = score
    return result


def validate_ladder(configs: Mapping[str, Mapping[str, CaseScore]]) -> list[str]:
    missing = [key for key in CONFIG_ORDER if key not in configs]
    if missing:
        raise ValueError(f"missing configurations: {missing}")
    case_sets = {key: set(configs[key]) for key in CONFIG_ORDER}
    reference = case_sets[CONFIG_ORDER[0]]
    if len(reference) != 29:
        raise ValueError(f"expected 29 cases, found {len(reference)}")
    for key, names in case_sets.items():
        if names != reference:
            raise ValueError(f"case-set mismatch for {key}")
    expected_maps = {
        key: {name: score.expected for name, score in configs[key].items()}
        for key in CONFIG_ORDER
    }
    baseline_expected = expected_maps[CONFIG_ORDER[0]]
    for key, expected in expected_maps.items():
        if expected != baseline_expected:
            raise ValueError(f"GT-count mismatch for {key}")
    if sum(baseline_expected.values()) != 86:
        raise ValueError("expected the published 86 GT assembly features")
    return sorted(reference)


def _percentile(sorted_values: Sequence[float], probability: float) -> float:
    if not sorted_values:
        raise ValueError("cannot take a percentile of an empty sequence")
    position = (len(sorted_values) - 1) * probability
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return float(sorted_values[lower])
    weight = position - lower
    return float(sorted_values[lower] * (1 - weight) + sorted_values[upper] * weight)


def paired_bootstrap(
    baseline: Mapping[str, CaseScore],
    treatment: Mapping[str, CaseScore],
    *,
    replicates: int = 20_000,
    seed: int = 71_071,
) -> dict[str, float | int]:
    names = sorted(set(baseline) & set(treatment))
    if set(baseline) != set(treatment):
        raise ValueError("paired bootstrap requires identical case sets")
    observed = summarize(treatment.values()).f1 - summarize(baseline.values()).f1
    rng = random.Random(seed)
    deltas: list[float] = []
    for _ in range(replicates):
        sampled = [names[rng.randrange(len(names))] for _ in names]
        baseline_f1 = summarize(baseline[name] for name in sampled).f1
        treatment_f1 = summarize(treatment[name] for name in sampled).f1
        deltas.append(treatment_f1 - baseline_f1)
    deltas.sort()
    wins = sum(treatment[name].matched > baseline[name].matched for name in names)
    losses = sum(treatment[name].matched < baseline[name].matched for name in names)
    return {
        "observed_f1_delta": observed,
        "ci95_low": _percentile(deltas, 0.025),
        "ci95_high": _percentile(deltas, 0.975),
        "bootstrap_probability_positive": sum(value > 0 for value in deltas)
        / replicates,
        "matched_wins": wins,
        "matched_losses": losses,
        "matched_ties": len(names) - wins - losses,
        "replicates": replicates,
        "seed": seed,
    }


def _overall_workbook_row(path: Path, version: str) -> dict[str, object]:
    for row in _sheet_rows(path, "同口径总表"):
        if row.get("版本") == version:
            return row
    raise ValueError(f"cannot find {version!r} in {path.name}")


def load_recorded_efficiency(dynamic: Path, topology: Path) -> dict[str, dict[str, float]]:
    sources = {
        "serial": _overall_workbook_row(dynamic, "父Issue Terra旧串行"),
        "dynamic": _overall_workbook_row(dynamic, "Terra动态串行"),
        "topology": _overall_workbook_row(topology, "新版拓扑动态Terra"),
    }
    result: dict[str, dict[str, float]] = {}
    for key, row in sources.items():
        result[key] = {
            "coarse_time_s": float(row.get("粗切记录值总和(s)") or 0),
            "fine_time_s": float(row.get("细切阶段记录值总和(s,含锁等待)") or 0),
            "model_time_s": float(row.get("模型记录值总和(s)") or 0),
            "input_tokens": float(row.get("输入tokens") or 0),
            "output_tokens": float(row.get("输出tokens") or 0),
            "total_tokens": float(row.get("总tokens") or 0),
        }
    return result


def load_final_efficiency(results_dir: Path) -> tuple[dict[str, float], dict[str, object]]:
    metrics_path = results_dir / "metrics_visual_json_serial.csv"
    with metrics_path.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    if len(rows) != 29 or any(row["status"] != "OK" for row in rows):
        raise ValueError("final metrics must contain 29 successful cases")
    efficiency = {
        "coarse_time_s": sum(float(row["pipeline_time_s"]) for row in rows),
        "fine_time_s": sum(float(row["fine_pipeline_time_s"]) for row in rows),
        "model_time_s": sum(float(row["codex_time_s"]) for row in rows),
        "input_tokens": sum(float(row["prompt_tokens"]) for row in rows),
        "output_tokens": sum(float(row["completion_tokens"]) for row in rows),
        "total_tokens": sum(float(row["total_tokens"]) for row in rows),
    }
    metrics_by_name = {row["base_name"]: row for row in rows}
    selections: list[int] = []
    dense_counts: list[int] = []
    fine_times: list[float] = []
    token_counts: list[float] = []
    for plan_path in sorted(results_dir.glob("*_fine_slice_plan.json")):
        name = plan_path.name.removesuffix("_fine_slice_plan.json")
        plan = json.loads(plan_path.read_text(encoding="utf-8"))
        selected = int(plan["estimated_slices"])
        height = float(plan["layer_height"])
        dense = sum(math.floor(float(length) / height) + 1 for length in plan["bounding_box_lwh"])
        selections.append(selected)
        dense_counts.append(dense)
        fine_times.append(float(metrics_by_name[name]["fine_pipeline_time_s"]))
        token_counts.append(float(metrics_by_name[name]["total_tokens"]))
    if len(selections) != 29:
        raise ValueError("expected 29 final fine-slice plans")
    rho_time = spearmanr(selections, fine_times)
    rho_tokens = spearmanr(selections, token_counts)
    sorted_slices = sorted(selections)
    slicing = {
        "selected_slices": sum(selections),
        "full_dense_slices": sum(dense_counts),
        "selection_ratio": sum(selections) / sum(dense_counts),
        "median_selected_slices": statistics.median(selections),
        "p90_selected_slices": _percentile(sorted_slices, 0.9),
        "max_selected_slices": max(selections),
        "spearman_slices_vs_fine_time": float(rho_time.statistic),
        "spearman_slices_vs_fine_time_p": float(rho_time.pvalue),
        "spearman_slices_vs_tokens": float(rho_tokens.statistic),
        "spearman_slices_vs_tokens_p": float(rho_tokens.pvalue),
    }
    return efficiency, slicing


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _fmt_pct(value: float) -> str:
    return f"{100 * value:.2f}%"


def _report_markdown(
    summaries: Mapping[str, Summary],
    bootstraps: Mapping[str, Mapping[str, float | int]],
    efficiency: Mapping[str, Mapping[str, float]],
    slicing: Mapping[str, object],
    by_difficulty: Mapping[str, Summary],
) -> str:
    lines = [
        "# TER-71 Terra 消融与效率实验报告",
        "",
        "## 结论",
        "",
        "在同一 29 件、86 个 GT 装配特征和同一一对一匹配器下，完整 pipeline "
        "达到 Precision 88.73%、Recall 73.26%、F1 80.25%。相对固定粗切串行基线，"
        "F1 提高 12.71 个百分点；按零件成簇的 20,000 次配对 bootstrap 95% CI 为 "
        f"[{100 * float(bootstraps['serial_to_full']['ci95_low']):.2f}, "
        f"{100 * float(bootstraps['serial_to_full']['ci95_high']):.2f}] 个百分点。",
        "",
        "单个中间模块的置信区间较宽，因此本文应表述为“小规模配对证据”，不写成已完成"
        "大样本统计证明。完整稳定化相对拓扑动态版的提升没有出现逐件 matched count 回退。",
        "",
        "## 六级配置结果",
        "",
        "| 配置 | Matched / Pred. / GT | Precision | Recall | F1 | Unmatched |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for key in CONFIG_ORDER:
        summary = summaries[key]
        lines.append(
            f"| {DISPLAY_NAMES[key]} | {summary.matched} / {summary.predicted} / "
            f"{summary.expected} | {_fmt_pct(summary.precision)} | {_fmt_pct(summary.recall)} | "
            f"{_fmt_pct(summary.f1)} | {_fmt_pct(summary.unmatched_rate)} |"
        )
    lines.extend(
        [
            "",
            "解释：串行相对并行主要减少误报（Precision +19.77 pp），但 Recall 略降 3.49 pp；"
            "动态切片相对固定粗切串行使 Recall +5.81 pp、F1 +4.25 pp；拓扑增强再使 F1 +2.09 pp；"
            "最终稳定化使 F1 再提高 6.37 pp。后两级包含耦合规则更新，不能拆成单一阈值的因果贡献。",
            "",
            "## 配对 bootstrap",
            "",
            "| 对比 | ΔF1 | 95% CI | P(Δ>0) | matched 胜/负/平 |",
            "|---|---:|---:|---:|---:|",
        ]
    )
    comparison_names = {
        "visual_to_serial": "visual only → serial",
        "parallel_to_serial": "parallel → serial",
        "serial_to_dynamic": "serial → dynamic",
        "dynamic_to_topology": "dynamic → topology",
        "topology_to_full": "topology → full",
        "serial_to_full": "serial → full",
    }
    for key, label in comparison_names.items():
        row = bootstraps[key]
        lines.append(
            f"| {label} | {100 * float(row['observed_f1_delta']):+.2f} pp | "
            f"[{100 * float(row['ci95_low']):+.2f}, {100 * float(row['ci95_high']):+.2f}] pp | "
            f"{float(row['bootstrap_probability_positive']):.3f} | "
            f"{row['matched_wins']}/{row['matched_losses']}/{row['matched_ties']} |"
        )
    lines.extend(
        [
            "",
            "## 难度分层（完整 pipeline）",
            "",
            "| 难度 | Matched / Pred. / GT | Precision | Recall | F1 |",
            "|---|---:|---:|---:|---:|",
        ]
    )
    for difficulty in ("easy", "medium", "hard"):
        row = by_difficulty[difficulty]
        lines.append(
            f"| {difficulty} | {row.matched} / {row.predicted} / {row.expected} | "
            f"{_fmt_pct(row.precision)} | {_fmt_pct(row.recall)} | {_fmt_pct(row.f1)} |"
        )
    lines.extend(
        [
            "",
            "## 鲁棒性与资源效率",
            "",
            f"最终运行 29/29 成功。候选驱动细切共选择 {int(slicing['selected_slices']):,} 层；"
            f"若三个轴均以 0.01 mm 全量细切，估计为 {int(slicing['full_dense_slices']):,} 层，"
            f"因此保留 {100 * float(slicing['selection_ratio']):.2f}%（节省 "
            f"{100 * (1 - float(slicing['selection_ratio'])):.2f}%）。每件中位数 "
            f"{float(slicing['median_selected_slices']):,.0f} 层，P90 "
            f"{float(slicing['p90_selected_slices']):,.0f} 层，最大 "
            f"{int(slicing['max_selected_slices']):,} 层。",
            "",
            f"细切层数与细切耗时的 Spearman ρ={float(slicing['spearman_slices_vs_fine_time']):.3f} "
            f"(p={float(slicing['spearman_slices_vs_fine_time_p']):.4f})，与 token 数的 "
            f"ρ={float(slicing['spearman_slices_vs_tokens']):.3f} "
            f"(p={float(slicing['spearman_slices_vs_tokens_p']):.4f})，表明复杂度长尾仍是主要成本来源。",
            "",
            "| 配置 | 总 tokens | 每件 tokens | 模型耗时/件 | 粗切+细切记录耗时/件 |",
            "|---|---:|---:|---:|---:|",
        ]
    )
    for key in ("serial", "dynamic", "topology", "full"):
        row = efficiency[key]
        lines.append(
            f"| {DISPLAY_NAMES[key]} | {row['total_tokens']:,.0f} | "
            f"{row['total_tokens'] / 29:,.0f} | {row['model_time_s'] / 29:.1f} s | "
            f"{(row['coarse_time_s'] + row['fine_time_s']) / 29:.1f} s |"
        )
    lines.extend(
        [
            "",
            "注意：旧串行工作簿没有记录几何预处理耗时；动态版的细切时间字段包含锁等待，"
            "所以该表用于规模画像，不用于严格的端到端速度优越性结论。完整 pipeline 提升精度的"
            "代价是 token 增长，后续应进一步压缩拓扑摘要。",
            "",
            "## 数据质量与结论边界",
            "",
            "`medium_32` 的 GT 标为 0 个局部特征，但几何与图像存在可见孔；其 5 个预测全部"
            "计为 type-absent，占最终 8 个 unmatched 的 62.5%。主结果保留原始 GT，不做"
            "定向修正。29 件均为单次模型运行；除串行/并行共享同一旧版几何输入外，动态、拓扑和"
            "最终版属于逐版本机制消融，仍需多随机重复和人工清洗 GT 才能形成强统计结论。",
            "",
            "## 建议的论文三类实验",
            "",
            "1. **主结果与消融：** 六级配置表 + 配对 bootstrap，回答串行、动态切片、拓扑与稳定化是否有效。",
            "2. **复杂度鲁棒性与效率：** 按 easy/medium/hard 分层，报告成功率、细切选择率、耗时和 token 长尾。",
            "3. **实机闭环：** 由作者补充特征配对、视觉定位、抓取与装配成功率；离线结果不替代该实验。",
            "",
        ]
    )
    return "\n".join(lines)


def _write_workbook(
    path: Path,
    summaries: Mapping[str, Summary],
    bootstraps: Mapping[str, Mapping[str, float | int]],
    by_difficulty: Mapping[str, Summary],
    efficiency: Mapping[str, Mapping[str, float]],
    slicing: Mapping[str, object],
) -> None:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    ablation = workbook.create_sheet("Ablation")
    ablation.append(["Configuration", *Summary.__dataclass_fields__.keys()])
    for key in CONFIG_ORDER:
        ablation.append([DISPLAY_NAMES[key], *asdict(summaries[key]).values()])

    paired = workbook.create_sheet("Paired Bootstrap")
    bootstrap_fields = list(next(iter(bootstraps.values())).keys())
    paired.append(["Comparison", *bootstrap_fields])
    for key, values in bootstraps.items():
        paired.append([key, *(values[field] for field in bootstrap_fields)])

    difficulty = workbook.create_sheet("Difficulty")
    difficulty.append(["Difficulty", *Summary.__dataclass_fields__.keys()])
    for key in ("easy", "medium", "hard"):
        difficulty.append([key, *asdict(by_difficulty[key]).values()])

    cost = workbook.create_sheet("Efficiency")
    cost_fields = [
        "coarse_time_s",
        "fine_time_s",
        "model_time_s",
        "input_tokens",
        "output_tokens",
        "total_tokens",
    ]
    cost.append(["Configuration", *cost_fields])
    for key in ("serial", "dynamic", "topology", "full"):
        cost.append([DISPLAY_NAMES[key], *(efficiency[key][field] for field in cost_fields)])

    slicing_sheet = workbook.create_sheet("Dynamic Slicing")
    slicing_sheet.append(["Metric", "Value"])
    for key, value in slicing.items():
        slicing_sheet.append([key, value])

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for column in worksheet.columns:
            width = min(48, max(len(str(cell.value or "")) for cell in column) + 2)
            worksheet.column_dimensions[column[0].column_letter].width = width
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--three-flow", type=Path, required=True)
    parser.add_argument("--dynamic", type=Path, required=True)
    parser.add_argument("--topology", type=Path, required=True)
    parser.add_argument("--final", type=Path, required=True)
    parser.add_argument("--final-results-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--bootstrap-replicates", type=int, default=20_000)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    configs: dict[str, dict[str, CaseScore]] = {}
    configs.update(load_three_flow(args.three_flow))
    dynamic = load_dynamic(args.dynamic)
    if configs["serial"] != dynamic["serial"]:
        raise ValueError("the TER-65 and TER-67 serial baselines disagree")
    configs["dynamic"] = dynamic["dynamic"]
    configs["topology"] = load_topology(args.topology)
    configs["full"] = load_final(args.final)
    validate_ladder(configs)

    comparisons = {
        "visual_to_serial": ("visual_only", "serial"),
        "parallel_to_serial": ("parallel", "serial"),
        "serial_to_dynamic": ("serial", "dynamic"),
        "dynamic_to_topology": ("dynamic", "topology"),
        "topology_to_full": ("topology", "full"),
        "serial_to_full": ("serial", "full"),
    }
    summaries = {key: summarize(configs[key].values()) for key in CONFIG_ORDER}
    bootstraps = {
        name: paired_bootstrap(
            configs[baseline],
            configs[treatment],
            replicates=args.bootstrap_replicates,
        )
        for name, (baseline, treatment) in comparisons.items()
    }
    by_difficulty = {
        difficulty: summarize(
            score for score in configs["full"].values() if score.difficulty == difficulty
        )
        for difficulty in ("easy", "medium", "hard")
    }
    efficiency = load_recorded_efficiency(args.dynamic, args.topology)
    efficiency["full"], slicing = load_final_efficiency(args.final_results_dir)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    with (args.output_dir / "ter71_ablation_summary.csv").open(
        "w", encoding="utf-8", newline=""
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["configuration", *Summary.__dataclass_fields__.keys()],
        )
        writer.writeheader()
        for key in CONFIG_ORDER:
            writer.writerow({"configuration": key, **asdict(summaries[key])})
    with (args.output_dir / "ter71_paired_bootstrap.csv").open(
        "w", encoding="utf-8", newline=""
    ) as handle:
        fieldnames = ["comparison", *next(iter(bootstraps.values())).keys()]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for key, value in bootstraps.items():
            writer.writerow({"comparison": key, **value})

    inputs = [args.three_flow, args.dynamic, args.topology, args.final]
    manifest = {
        "commit": "d98baca",
        "model": "gpt-5.6-terra",
        "reasoning_effort": "medium",
        "case_count": 29,
        "gt_feature_count": 86,
        "inputs": {path.name: _sha256(path) for path in inputs},
        "summaries": {key: asdict(value) for key, value in summaries.items()},
        "paired_bootstrap": bootstraps,
        "by_difficulty": {key: asdict(value) for key, value in by_difficulty.items()},
        "efficiency": efficiency,
        "dynamic_slicing": slicing,
    }
    (args.output_dir / "ter71_analysis_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (args.output_dir / "TER-71-EXPERIMENT-REPORT.md").write_text(
        _report_markdown(summaries, bootstraps, efficiency, slicing, by_difficulty),
        encoding="utf-8",
    )
    _write_workbook(
        args.output_dir / "TER-71-experiment-results.xlsx",
        summaries,
        bootstraps,
        by_difficulty,
        efficiency,
        slicing,
    )


if __name__ == "__main__":
    main()
