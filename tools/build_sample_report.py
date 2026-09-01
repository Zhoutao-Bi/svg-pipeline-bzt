"""Build a four-row comparison workbook for one three-workflow sample run."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
MODES = [
    ("visual_only", "纯视觉"),
    ("visual_json_parallel", "视觉+JSON并行"),
    ("visual_json_serial", "视觉→JSON串行矫正"),
]
TARGET_TYPES = {"孔", "柱", "槽", "倒角"}


def number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    found = re.search(r"[-+]?\d+(?:\.\d+)?", str(value))
    return float(found.group()) if found else None


def normalize_gt_type(label: str | None) -> str | None:
    text = str(label or "")
    if "孔" in text:
        return "孔"
    if "槽" in text:
        return "槽"
    if "倒角" in text or "圆角" in text:
        return "倒角"
    if any(word in text for word in ("柱", "凸台", "外环", "主直径", "收细段", "半圆")):
        return "柱"
    return None


def load_ground_truth(path: Path, sample: str) -> dict:
    ws = load_workbook(path, read_only=True, data_only=True)["Sheet2"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0] or "") != sample:
            continue
        features = []
        for index in range(5, len(row), 5):
            if row[index] is None:
                continue
            features.append(
                {
                    "原始名称": str(row[index]),
                    "特征类型": normalize_gt_type(row[index]),
                    "特征形状": None,
                    "坐标X": number(row[index + 1]),
                    "坐标Y": number(row[index + 2]),
                    "坐标Z": number(row[index + 3]),
                    "尺寸类型": None,
                    "尺寸数据": number(row[index + 4]),
                    "作用": "装配特征",
                }
            )
        return {
            "名字": sample,
            "整体特征": row[1],
            "尺寸X": number(row[2]),
            "尺寸Y": number(row[3]),
            "尺寸Z": number(row[4]),
            "局部特征列表": features,
        }
    raise ValueError(f"参考答案中找不到样本: {sample}")


def rmse(values: list[float]) -> float | None:
    return math.sqrt(sum(value * value for value in values) / len(values)) if values else None


def evaluate(prediction: dict, ground_truth: dict) -> dict:
    pred_features = [
        feature
        for feature in prediction.get("局部特征列表", [])
        if feature.get("特征类型") in TARGET_TYPES and feature.get("作用") == "装配特征"
    ]
    gt_features = [
        feature
        for feature in ground_truth.get("局部特征列表", [])
        if feature.get("特征类型") in TARGET_TYPES
    ]
    dimensions = [number(ground_truth.get(key)) or 0.0 for key in ("尺寸X", "尺寸Y", "尺寸Z")]
    diagonal = max(math.sqrt(sum(value * value for value in dimensions)), 1.0)
    candidates = []
    for pred_index, pred_feature in enumerate(pred_features):
        for gt_index, gt_feature in enumerate(gt_features):
            if pred_feature.get("特征类型") != gt_feature.get("特征类型"):
                continue
            pred_coords = [number(pred_feature.get(key)) for key in ("坐标X", "坐标Y", "坐标Z")]
            gt_coords = [number(gt_feature.get(key)) for key in ("坐标X", "坐标Y", "坐标Z")]
            if all(value is not None for value in pred_coords + gt_coords):
                distance = math.sqrt(sum((pred - gt) ** 2 for pred, gt in zip(pred_coords, gt_coords)))
            else:
                distance = None
            pred_size = number(pred_feature.get("尺寸数据"))
            gt_size = number(gt_feature.get("尺寸数据"))
            size_error = abs(pred_size - gt_size) if pred_size is not None and gt_size is not None else None
            relative_size_error = size_error / max(abs(gt_size), 1.0) if size_error is not None else 0.0
            normalized_distance = distance / diagonal if distance is not None else 0.0
            if normalized_distance <= 0.20 and relative_size_error <= 1.0:
                candidates.append(
                    (normalized_distance + 0.25 * relative_size_error, pred_index, gt_index, distance, size_error)
                )

    used_pred, used_gt, matches = set(), set(), []
    for _, pred_index, gt_index, distance, size_error in sorted(candidates):
        if pred_index in used_pred or gt_index in used_gt:
            continue
        used_pred.add(pred_index)
        used_gt.add(gt_index)
        matches.append((distance, size_error))

    dimension_errors = []
    for key in ("尺寸X", "尺寸Y", "尺寸Z"):
        pred_value, gt_value = number(prediction.get(key)), number(ground_truth.get(key))
        if pred_value is not None and gt_value is not None:
            dimension_errors.append(pred_value - gt_value)

    return {
        "预测装配特征数": len(pred_features),
        "GT装配特征数": len(gt_features),
        "自动匹配数": len(matches),
        "未匹配预测数": len(pred_features) - len(matches),
        "漏检数": len(gt_features) - len(matches),
        "整体尺寸RMSE": rmse(dimension_errors),
        "坐标RMSE": rmse([distance for distance, _ in matches if distance is not None]),
        "特征尺寸RMSE": rmse([error for _, error in matches if error is not None]),
    }


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, sheet.max_column + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(1, sheet.max_row + 1)]
        sheet.column_dimensions[get_column_letter(column)].width = min(max(max(map(len, values)) + 2, 10), 32)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--results-dir", type=Path, required=True)
    parser.add_argument("--sample", required=True)
    parser.add_argument("--model-tag", required=True)
    parser.add_argument("--ground-truth", type=Path, default=ROOT / "result" / "old" / "exp1.xlsx")
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()

    results_dir = args.results_dir.resolve()
    output = (args.output or results_dir / f"{args.sample}_{args.model_tag}_三类流程_参考答案.xlsx").resolve()
    ground_truth = load_ground_truth(args.ground_truth.resolve(), args.sample)
    predictions = []
    for mode_key, mode_name in MODES:
        path = results_dir / f"{args.sample}_refined_{args.model_tag}_{mode_key}.txt"
        with path.open(encoding="utf-8") as handle:
            prediction = json.load(handle)
        prediction.update({"名字": args.sample, "流程": mode_name, "流程键": mode_key, "结果文件": path.name})
        predictions.append(prediction)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "三类流程+参考答案"
    headers = [
        "文件名", "行类型", "流程", "模型", "整体特征", "尺寸X", "尺寸Y", "尺寸Z",
        "局部特征总数", "预测装配特征数", "GT装配特征数", "自动匹配数", "未匹配预测数",
        "漏检数", "整体尺寸RMSE", "坐标RMSE", "特征尺寸RMSE", "局部特征JSON", "结果文件",
    ]
    sheet.append(headers)
    for prediction in predictions:
        metrics = evaluate(prediction, ground_truth)
        sheet.append(
            [
                args.sample, "预测", prediction["流程"], args.model_tag, prediction.get("整体特征"),
                prediction.get("尺寸X"), prediction.get("尺寸Y"), prediction.get("尺寸Z"),
                len(prediction.get("局部特征列表", [])),
                *[metrics[key] for key in headers[9:17]],
                json.dumps(prediction.get("局部特征列表", []), ensure_ascii=False), prediction["结果文件"],
            ]
        )
    sheet.append(
        [
            args.sample, "参考答案", "参考答案", "GT", ground_truth.get("整体特征"),
            ground_truth.get("尺寸X"), ground_truth.get("尺寸Y"), ground_truth.get("尺寸Z"),
            len(ground_truth.get("局部特征列表", [])), None, len(ground_truth.get("局部特征列表", [])),
            None, None, None, None, None, None,
            json.dumps(ground_truth.get("局部特征列表", []), ensure_ascii=False), args.ground_truth.name,
        ]
    )
    style_sheet(sheet)
    group_border = Border(bottom=Side(style="medium", color="5B9BD5"))
    for cell in sheet[sheet.max_row]:
        cell.border = group_border

    metrics_sheet = workbook.create_sheet("运行指标")
    metric_headers = [
        "base_name", "流程", "pipeline_time_s", "codex_time_s", "prompt_tokens",
        "completion_tokens", "total_tokens", "status", "error",
    ]
    metrics_sheet.append(metric_headers)
    for mode_key, mode_name in MODES:
        with (results_dir / f"metrics_{mode_key}.csv").open(encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                metrics_sheet.append([row.get("base_name"), mode_name] + [row.get(key) for key in metric_headers[2:]])
    style_sheet(metrics_sheet)

    notes_sheet = workbook.create_sheet("说明")
    notes_sheet.append(["项目", "说明"])
    notes_sheet.append(["实验模型", f"{args.model_tag}，medium reasoning，Codex ChatGPT/OAuth"])
    notes_sheet.append(["样本", f"仅 {args.sample} 一件，三类流程；串行流程包含两次模型调用"])
    notes_sheet.append(["四行结构", "纯视觉、视觉+JSON并行、视觉→JSON串行矫正、参考答案"])
    notes_sheet.append(["自动匹配", "同类型、坐标误差≤GT包络对角线20%、尺寸相对误差≤100%的一对一最小代价匹配"])
    notes_sheet.append(["指标边界", "“未匹配预测数”不是严格幻觉数；其中可能包含真实但定位/尺寸不合格、重复预测或类型错误，需人工复核"])
    notes_sheet.append(["结论边界", "单个 easy 样本只用于查看输出，不代表完整数据集模型结论"])
    style_sheet(notes_sheet)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    print(output)


if __name__ == "__main__":
    main()
